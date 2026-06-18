"""
Generates docs/Planwise Data Migration Template.xlsx — a workbook with one
sheet per phase from DATA_MIGRATION_TEMPLATE.md. Header cells are color-
coded (red = required, amber = recommended, grey = optional) and carry
tooltip comments. Each sheet has one example row in grey italics that
the customer is expected to delete.

Run from repo root:
    python docs/_build_migration_template.py

The generator script lives at docs/_build_migration_template.py and is
kept under version control so the workbook can be regenerated reliably.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\yulia\Planwise\docs\Planwise Data Migration Template.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)

H_REQUIRED = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
H_RECOMM = Font(name="Calibri", size=11, bold=True, color="000000")
H_OPTIONAL = Font(name="Calibri", size=11, italic=True, color="4B4B4B")
F_REQUIRED = PatternFill("solid", fgColor="C0392B")
F_RECOMM = PatternFill("solid", fgColor="F4D03F")
F_OPTIONAL = PatternFill("solid", fgColor="ECF0F1")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(cell, kind):
    if kind == "R":
        cell.font, cell.fill = H_REQUIRED, F_REQUIRED
    elif kind == "r":
        cell.font, cell.fill = H_RECOMM, F_RECOMM
    else:
        cell.font, cell.fill = H_OPTIONAL, F_OPTIONAL
    cell.alignment, cell.border = CENTER, BORDER


def add_sheet(name, columns, intro_lines):
    """columns: list of (name, kind, comment, sample_value)."""
    ws = wb.create_sheet(name)
    for i, line in enumerate(intro_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(italic=True, color="555555")
        c.alignment = LEFT
        ws.merge_cells(
            start_row=i, start_column=1, end_row=i, end_column=max(len(columns), 4)
        )
    header_row = len(intro_lines) + 2
    example_row_num = header_row + 1
    for col_idx, (col, kind, comment, sample) in enumerate(columns, start=1):
        h = ws.cell(row=header_row, column=col_idx, value=col)
        style_header(h, kind)
        if comment:
            h.comment = Comment(comment, "Planwise")
        e = ws.cell(row=example_row_num, column=col_idx, value=sample)
        e.alignment = LEFT
        e.border = BORDER
        e.font = Font(italic=True, color="707070")
        width = min(max(len(col), len(str(sample or "")) + 2, 14), 36)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(row=example_row_num + 1, column=1)
    return ws


# ── ReadMe ────────────────────────────────────────────────────────────
ws = wb.create_sheet("ReadMe")
APO = "’"
DONT = "do" + APO + "t"
lines = [
    ("Planwise — Customer Data Migration Workbook",
     Font(size=16, bold=True, color="1F4E79")),
    ("", None),
    ("Fill in one tab at a time, in phase order. Each tab has one example row in grey italics — DELETE it before saving.", None),
    ("", None),
    ("Header colors:", Font(bold=True)),
    ("  RED   = Required. Row is rejected if blank.", Font(color="C0392B", bold=True)),
    ("  AMBER = Recommended. Row is accepted, but importer flags it.", Font(color="B7950B", bold=True)),
    ("  GREY  = Optional. Leave blank if you " + DONT + " have it.", Font(italic=True, color="4B4B4B")),
    ("", None),
    ("Hover over any header cell for column-specific notes.", None),
    ("", None),
    ("Recommended order (one wave per phase — no need to do them all at once):", Font(bold=True)),
    ("  Phase 1 — Foundation:        Users, Customers, Catalogs", None),
    ("  Phase 2 — Active projects:    Projects, Tasks (+ one XML file per project for zones)", None),
    ("  Phase 3 — Operations:         Assignments, Timesheet", None),
    ("", None),
    ("Full spec: docs/DATA_MIGRATION_TEMPLATE.md in the Planwise repo.",
     Font(italic=True, color="555555")),
]
for i, (text, font) in enumerate(lines, start=1):
    c = ws.cell(row=i, column=1, value=text)
    if font is not None:
        c.font = font
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
ws.column_dimensions["A"].width = 110

# ── Phase 1a — Users ──────────────────────────────────────────────────
add_sheet(
    "1a Users",
    columns=[
        ("first_name", "R", "Required. Login user given name.", "Yossi"),
        ("last_name", "R", "Required.", "Cohen"),
        ("email", "R", "Required. Unique login identifier.", "yossi@example.com"),
        ("phone", "r", "Recommended.", "+972-50-1234567"),
        ("job_title", "r", "Recommended. Must match the JobTitles catalog (sheet 1c JobTitles). Drives the role-picker filter.", "BIM Coordinator"),
        ("department", "r", "Recommended.", "MEP Coordination"),
        ("seniority", "o", "Optional. Junior / Mid / Senior / Lead / Professional employee.", "Senior"),
        ("is_active", "o", "Optional. true (default) or false.", "true"),
        ("code", "o", "Optional. Internal employee number.", "EMP-00000103"),
    ],
    intro_lines=["Phase 1a — Users (people who will log in)."],
)

# ── Phase 1b — Customers ──────────────────────────────────────────────
add_sheet(
    "1b Customers",
    columns=[
        ("company_name", "R", "Required. Customer organisation name.", "Municipality of Tel Aviv"),
        ("tax_id", "r", "Recommended. Tax / company number — strongest dedupe key.", "514123456"),
        ("email", "r", "Recommended. Office email — also a dedupe key.", "office@tlv.gov.example"),
        ("phone", "r", "Recommended. Main switchboard.", "+972-3-1234567"),
        ("address", "r", "Recommended. One free-text line.", "Ibn Gvirol 69, Tel Aviv"),
        ("website", "o", "Optional.", "https://www.tlv.gov.example"),
        ("notes", "o", "Optional. Free text.", ""),
    ],
    intro_lines=[
        "Phase 1b — Customers (organisations the company serves).",
        "Importer auto-fills partner_type=organization and main_role_type_code=customer.",
    ],
)

# ── Phase 1c — Catalogs ───────────────────────────────────────────────
add_sheet(
    "1c Phases",
    columns=[
        ("name", "R", "Required. Phase display name (Hebrew or English).", "תכנון מפורט"),
        ("sort_order", "o", "Optional. Integer — controls display order.", "21"),
    ],
    intro_lines=[
        "Phase 1c — Phases catalog. Reuse our defaults and add any customer-specific stages."
    ],
)
add_sheet(
    "1c ProjectTypes",
    columns=[("name", "R", "Required. Display name.", "BIM Coordination")],
    intro_lines=["Phase 1c — Project types catalog."],
)
add_sheet(
    "1c Services",
    columns=[
        ("department", "R", "Required. Department this service belongs to.", "תיאום מערכות"),
        ("service_name", "R", "Required.", "תיאום מערכות BIM"),
        ("deliverable_codes", "o", "Optional. Comma-separated DLV codes from sheet 1c Deliverables.", "DLV-101,DLV-102,DLV-103"),
    ],
    intro_lines=["Phase 1c — Services catalog (service lines the company offers)."],
)
add_sheet(
    "1c Deliverables",
    columns=[
        ("service", "R", "Required. Service this deliverable belongs to.", "תיאום מערכות"),
        ("deliverable_name", "R", "Required.", "דוח קריטי"),
        ("deliverable_code", "r", "Recommended. DLV-NNN — unique. Used as the join key.", "DLV-101"),
        ("typical_task_codes", "o", "Optional. Comma-separated TSK codes that typically belong to this deliverable.", "TSK-1062,TSK-1065"),
    ],
    intro_lines=["Phase 1c — Deliverables catalog (each service has one or more deliverables)."],
)
add_sheet(
    "1c JobTitles",
    columns=[
        ("name", "R", "Required. Used by project role-pickers to filter eligible people.", "BIM Coordinator"),
        ("sort_order", "o", "Optional.", "10"),
    ],
    intro_lines=["Phase 1c — Job Titles catalog. Drives the Job-Title filter on role pickers."],
)

# ── Phase 2a — Projects ───────────────────────────────────────────────
add_sheet(
    "2a Projects",
    columns=[
        ("project_name", "R", "Required.", "דרך חברון אסדן"),
        ("project_number", "R", "Required. Unique internal number.", "260066"),
        ("customer", "R", "Required. Must match a company_name from sheet 1b Customers.", "אסדן"),
        ("project_type", "r", "Recommended. Comma-separated values from sheet 1c ProjectTypes.", "תיאום מערכות, ניהול מודל"),
        ("phase", "r", "Recommended. Must match a name from sheet 1c Phases.", "תכנון סופי"),
        ("status", "r", "Recommended. active / on_hold / completed / cancelled. Default active.", "active"),
        ("start_date", "r", "Recommended. DD.MM.YYYY or ISO.", "01.06.2026"),
        ("end_date", "r", "Recommended.", "12.12.2027"),
        ("contract_amount", "r", "Recommended. Numeric — currency defaults to ILS.", "289000"),
        ("pm", "r", "Recommended. Project Manager — must match a person from sheet 1a Users.", "Shiffy Klaynman"),
        ("weekly_meeting_day", "o", "Optional. Free text.", "monday at 12:00"),
        ("authoring_tool_version", "o", "Optional.", "R24"),
        ("services_per_contract", "o", "Optional. Free-text description of the contract scope.", "2 towers + commercial building over shared basement"),
        ("notes", "o", "Optional.", ""),
    ],
    intro_lines=[
        "Phase 2a — Projects (header rows).",
        "Zone breakdown for each project lives in a SEPARATE XML file — see docs/DATA_MIGRATION_TEMPLATE.md §2b.",
    ],
)

# ── Phase 2c — Tasks ──────────────────────────────────────────────────
add_sheet(
    "2c Tasks",
    columns=[
        ("project_number", "R", "Required. Must match sheet 2a Projects.", "260066"),
        ("zone_id", "r", "Recommended. ID from the project XML, or NA for project-level.", "B0031"),
        ("task_name", "R", "Required.", "הקמת דוח קריטי"),
        ("task_code", "r", "Recommended. TSK-NNNN from the Typical_Tasks catalog OR custom.", "TSK-1062"),
        ("service", "r", "Recommended. Must match sheet 1c Services.", "תיאום מערכות BIM"),
        ("deliverable", "r", "Recommended. Deliverable name OR DLV-NNN from sheet 1c Deliverables.", "DLV-101"),
        ("budget_hours", "r", "Recommended. Numeric.", "8"),
        ("start_date", "o", "Optional.", ""),
        ("end_date", "o", "Optional.", "23.06.2026"),
        ("status", "o", "Optional. Default not_started.", "not_started"),
        ("priority", "o", "Optional. low / medium / high / critical.", "medium"),
    ],
    intro_lines=["Phase 2c — Tasks per project."],
)

# ── Phase 3a — Assignments ────────────────────────────────────────────
add_sheet(
    "3a Assignments",
    columns=[
        ("project_number", "R", "Required.", "260066"),
        ("task_code", "R", "Required. Match the task by code.", "TSK-1062"),
        ("user_email", "R", "Required. Must match sheet 1a Users.", "yossi@example.com"),
        ("role_in_task", "o", "Optional. Reviewer / Lead / etc.", "Lead"),
        ("status", "o", "Optional. Status of THIS assignment.", "in_progress"),
    ],
    intro_lines=["Phase 3a — Who works on what (per-task assignments)."],
)

# ── Phase 3b — Timesheet ──────────────────────────────────────────────
add_sheet(
    "3b Timesheet",
    columns=[
        ("user_email", "R", "Required.", "yossi@example.com"),
        ("project_number", "R", "Required.", "260066"),
        ("task_code", "r", "Recommended. If known; otherwise a project-level entry.", "TSK-1062"),
        ("date", "R", "Required. DD.MM.YYYY.", "17.06.2026"),
        ("start_time", "r", "Recommended. HH:MM (24h). Provide either start+end OR minutes.", "09:00"),
        ("end_time", "r", "Recommended.", "12:00"),
        ("minutes", "o", "Optional. Total minutes — used if start/end not provided.", "180"),
        ("location", "o", "Optional. office / home.", "office"),
        ("is_billable", "o", "Optional. Default true.", "true"),
        ("notes", "o", "Optional.", "Critical report draft"),
    ],
    intro_lines=["Phase 3b — Historical timesheet entries."],
)

wb.save(OUT)
print("Wrote", OUT)
print("Sheets:", wb.sheetnames)
